# ============================================================
# Imports
# ============================================================
import io
import csv
import secrets
import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.db.models import Q
from django.db import transaction
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated
from rest_framework.decorators import action, api_view, permission_classes, authentication_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import (
    Course, Lecturer, Venue, StudentGroup, TimeSlot, ScheduleSlot,
    LecturerRequest, LecturerNotification, Announcement, StudentNotification,
    UserProfile,
)
from .serializers import (
    CourseSerializer, LecturerSerializer, VenueSerializer,
    StudentGroupSerializer, TimeSlotSerializer,
    ScheduleSlotReadSerializer, ScheduleSlotWriteSerializer,
    AuthUserSerializer,
    StudentProfileSerializer, LecturerProfileSerializer, LecturerRequestSerializer, LecturerNotificationSerializer,
    StudentAccountSerializer,
    AnnouncementSerializer, StudentNotificationSerializer,
)
from .emails import send_class_change_email, send_credentials_email
from .scheduler import generate_timetable_for_group

# ── PDF export ───────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
TIMES = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00',
         '14:00', '15:00', '16:00', '17:00']


# ============================================================
# Shared helpers
# ============================================================

def parse_uploaded_file(file):
    """
    Returns (fieldnames, list_of_row_dicts) — works for both .xlsx and .csv,
    so the rest of each bulk upload view's per-row logic never needs to
    care which format was actually uploaded.
    """
    name = (file.name or '').lower()

    if name.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else '' for h in next(rows_iter)]
        data_rows = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue  # skip fully blank rows
            row_dict = {
                headers[i]: ('' if row[i] is None else str(row[i]).strip())
                for i in range(len(headers)) if i < len(row)
            }
            data_rows.append(row_dict)
        return headers, data_rows

    # Fall back to CSV
    decoded = file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    return (reader.fieldnames or []), list(reader)


def generate_lecturer_id():
    year = datetime.datetime.now().year
    prefix = f"LEC-{year}-"
    lecturers_list = Lecturer.objects.filter(lecturer_id__startswith=prefix)
    max_num = 0
    for lec in lecturers_list:
        if lec.lecturer_id:
            try:
                parts = lec.lecturer_id.split('-')
                if len(parts) == 3:
                    num = int(parts[2])
                    if num > max_num:
                        max_num = num
            except (ValueError, IndexError):
                pass
    return f"{prefix}{max_num + 1:03d}"


def generate_random_password():
    chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789!@#$%^&*"
    return "".join(secrets.choice(chars) for _ in range(10))


class IsAdminRole(BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False

        profile = getattr(request.user, 'profile', None)
        return bool(request.user.is_staff or request.user.is_superuser or (profile and profile.role == 'ADMIN'))


def serialize_user(user, request=None):
    profile = getattr(user, 'profile', None)
    payload = {
        'id': user.id,
        'username': user.username,
        'role': profile.role if profile else 'ADMIN',
        'must_change_password': profile.must_change_password if profile else False,
    }
    if profile and profile.lecturer_id:
        payload['lecturer_id'] = profile.lecturer_id
    if profile and profile.student_group_id:
        payload['student_group_id'] = profile.student_group_id
    if profile:
        payload['name'] = user.get_full_name().strip() or user.username
        payload['email'] = user.email
        payload['contact_number'] = profile.contact_number or ''
        payload['registration_number'] = profile.registration_number or ''
        payload['avatar_url'] = request.build_absolute_uri(profile.avatar.url) if request and profile.avatar else None
    return payload


def get_student_profile(request):
    profile = getattr(request.user, 'profile', None)
    if profile is None or profile.role != 'STUDENT':
        return None
    return profile


def build_student_dashboard(profile, request, semester='S2-2026'):
    slots = ScheduleSlot.objects.select_related('timeslot', 'course', 'lecturer', 'venue', 'group').filter(group=profile.student_group, is_published=True)
    if semester:
        slots = slots.filter(semester=semester)
    slots = list(slots)

    now = timezone.localtime()
    today_name = now.strftime('%A')
    current_time = now.time()

    todays_slots = [slot for slot in slots if slot.timeslot.day == today_name]
    todays_remaining = [slot for slot in todays_slots if slot.timeslot.end_time >= current_time]
    todays_remaining.sort(key=lambda slot: slot.timeslot.start_time)

    total_minutes = 0
    for slot in slots:
        start_minutes = slot.timeslot.start_time.hour * 60 + slot.timeslot.start_time.minute
        end_minutes = slot.timeslot.end_time.hour * 60 + slot.timeslot.end_time.minute
        total_minutes += max(0, end_minutes - start_minutes)

    curriculum_courses = list(profile.student_group.courses.all()) if profile.student_group else []

    announcements = Announcement.objects.select_related('student_group').filter(
        Q(audience='FACULTY') |
        Q(audience='BATCH', student_group=profile.student_group) |
        Q(audience='GROUP', student_group=profile.student_group)
    ).order_by('-published_at')[:10]

    notifications = StudentNotification.objects.select_related(
        'student_group', 'schedule_slot', 'schedule_slot__course', 'schedule_slot__timeslot', 'schedule_slot__venue'
    ).filter(student_group=profile.student_group).order_by('-created_at')[:10]

    profile_data = StudentProfileSerializer(profile, context={'request': request}).data
    profile_data['enrolled_subjects'] = CourseSerializer(curriculum_courses, many=True).data

    return {
        'profile': profile_data,
        'stats': {
            'classes_today': len(todays_slots),
            'weekly_hours': round(total_minutes / 60, 1),
            'subjects_enrolled': len(curriculum_courses),
        },
        'timetable': ScheduleSlotReadSerializer(slots, many=True).data,
        'todays_remaining': ScheduleSlotReadSerializer(todays_remaining, many=True).data,
        'notifications': StudentNotificationSerializer(notifications, many=True).data,
        'announcements': AnnouncementSerializer(announcements, many=True).data,
        'today_name': today_name,
    }


# ============================================================
# Auth
# ============================================================

@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def auth_login(request):
    username = str(request.data.get('username', '')).strip()
    password = str(request.data.get('password', ''))

    user = authenticate(request, username=username, password=password)
    if not user:
        return Response({'detail': 'Invalid credentials'}, status=status.HTTP_400_BAD_REQUEST)

    profile = getattr(user, 'profile', None)
    actual_role = profile.role if profile else 'ADMIN'

    refresh = RefreshToken.for_user(user)

    return Response({
        'token': str(refresh.access_token),
        'refresh': str(refresh),
        'role': actual_role,
        'user': serialize_user(user, request=request),
    })


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        current_password = request.data.get('current_password', '')
        new_password = request.data.get('new_password', '')

        user = request.user
        if not user.check_password(current_password):
            return Response({'detail': 'Incorrect current password.'}, status=status.HTTP_400_BAD_REQUEST)

        if len(new_password) < 8:
            return Response({'detail': 'New password must be at least 8 characters long.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            user.set_password(new_password)
            user.save()

            profile = getattr(user, 'profile', None)
            if profile:
                profile.must_change_password = False
                profile.save(update_fields=['must_change_password'])

                if profile.lecturer:
                    profile.lecturer.must_change_password = False
                    profile.lecturer.save(update_fields=['must_change_password'])

        return Response({'detail': 'Password changed successfully.'}, status=status.HTTP_200_OK)


# ============================================================
# Core CRUD ViewSets
# ============================================================

class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer


class LecturerViewSet(viewsets.ModelViewSet):
    queryset = Lecturer.objects.all()
    serializer_class = LecturerSerializer


class StudentAccountViewSet(viewsets.ModelViewSet):
    serializer_class = StudentAccountSerializer
    permission_classes = [IsAdminRole]

    def get_queryset(self):
        return User.objects.select_related('profile', 'profile__student_group').filter(profile__role='STUDENT').order_by('username')


class LecturerMeViewSet(viewsets.ViewSet):
    def list(self, request):
        profile = getattr(request.user, 'profile', None)
        lecturer = profile.lecturer if profile else None
        if lecturer is None:
            return Response({'detail': 'Lecturer profile not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response(LecturerProfileSerializer(lecturer).data)

    def update(self, request):
        profile = getattr(request.user, 'profile', None)
        lecturer = profile.lecturer if profile else None
        if lecturer is None:
            return Response({'detail': 'Lecturer profile not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = LecturerProfileSerializer(lecturer, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venue.objects.all()
    serializer_class = VenueSerializer


class StudentGroupViewSet(viewsets.ModelViewSet):
    queryset = StudentGroup.objects.all()
    serializer_class = StudentGroupSerializer


class TimeSlotViewSet(viewsets.ModelViewSet):
    queryset = TimeSlot.objects.all()
    serializer_class = TimeSlotSerializer


class ScheduleSlotViewSet(viewsets.ModelViewSet):
    queryset = ScheduleSlot.objects.select_related(
        'timeslot', 'course', 'lecturer', 'venue', 'group'
    ).all()

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return ScheduleSlotWriteSerializer
        return ScheduleSlotReadSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params

        if level := params.get('level'):
            qs = qs.filter(group__level=level)
        if stream := params.get('stream'):
            qs = qs.filter(group__stream=stream)
        if day := params.get('day'):
            qs = qs.filter(timeslot__day=day)
        if semester := params.get('semester'):
            qs = qs.filter(semester=semester)
        if lecturer := params.get('lecturer'):
            qs = qs.filter(lecturer_id=lecturer)
        if group := params.get('group'):
            qs = qs.filter(group_id=group)

        return qs

    def perform_update(self, serializer):
        instance = self.get_object()
        old_timeslot_id = instance.timeslot_id
        old_venue_id = instance.venue_id
        was_published = instance.is_published

        updated = serializer.save()

        timeslot_changed = updated.timeslot_id != old_timeslot_id
        venue_changed = updated.venue_id != old_venue_id

        if not was_published:
            return
        if not (timeslot_changed or venue_changed):
            return  # nothing schedule-relevant changed, skip notifications

        student_notif_type = 'ROOM_CHANGE' if (venue_changed and not timeslot_changed) else 'RESCHEDULE'
        lecturer_notif_type = 'CHANGE'

        title = f"Schedule Updated: {updated.course.code}"
        message = (
            f"Your class has changed to {updated.timeslot.day} "
            f"{updated.timeslot.start_time.strftime('%H:%M')}-{updated.timeslot.end_time.strftime('%H:%M')} "
            f"at {updated.venue.code}."
        )

        LecturerNotification.objects.create(
            lecturer=updated.lecturer,
            schedule_slot=updated,
            title=title,
            message=message,
            notification_type=lecturer_notif_type,
        )
        StudentNotification.objects.create(
            student_group=updated.group,
            schedule_slot=updated,
            title=title,
            message=message,
            notification_type=student_notif_type,
        )

        student_emails = (
            User.objects
            .filter(profile__role='STUDENT', profile__student_group=updated.group)
            .exclude(email='')
            .values_list('email', flat=True)
        )

        send_class_change_email(
            notification_type=student_notif_type,
            course_code=updated.course.code,
            venue_code=updated.venue.code,
            day=updated.timeslot.day,
            start_time=updated.timeslot.start_time.strftime('%H:%M'),
            end_time=updated.timeslot.end_time.strftime('%H:%M'),
            lecturer_email=updated.lecturer.email,
            student_emails=student_emails,
        )

    def perform_destroy(self, instance):
        if instance.is_published:
            title = f"Cancelled: {instance.course.code}"
            message = (
                f"Your class on {instance.timeslot.day} "
                f"{instance.timeslot.start_time.strftime('%H:%M')} at {instance.venue.code} "
                f"has been cancelled."
            )
 
            LecturerNotification.objects.create(
                lecturer=instance.lecturer,
                title=title,
                message=message,
                notification_type='CANCEL',
            )
            StudentNotification.objects.create(
                student_group=instance.group,
                title=title,
                message=message,
                notification_type='CANCEL',
            )
 
            student_emails = (
                User.objects
                .filter(profile__role='STUDENT', profile__student_group=instance.group)
                .exclude(email='')
                .values_list('email', flat=True)
            )
 
            send_class_change_email(
                notification_type='CANCEL',
                course_code=instance.course.code,
                venue_code=instance.venue.code,
                day=instance.timeslot.day,
                start_time=instance.timeslot.start_time.strftime('%H:%M'),
                end_time=instance.timeslot.end_time.strftime('%H:%M'),
                lecturer_email=instance.lecturer.email,
                student_emails=student_emails,
            )
 
        instance.delete()

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        level = request.query_params.get('level', 'I')
        stream = request.query_params.get('stream', 'physical')
        semester = request.query_params.get('semester', 'S2-2026')
 
        slots = ScheduleSlot.objects.select_related(
            'timeslot', 'course', 'lecturer', 'venue', 'group'
        ).filter(
            group__level=level,
            group__stream=stream,
            semester=semester,
        )
 
        grid = {d: {} for d in DAYS}
        for slot in slots:
            hour = slot.timeslot.start_time.strftime('%H:%M')
            grid[slot.timeslot.day][hour] = slot
 
        # ---- Colour palette matching the web app's navy + brass identity ----
        navy = colors.HexColor('#0D1B2A')
        navy_soft = colors.HexColor('#16293F')
        brass = colors.HexColor('#C6963C')
        brass_light = colors.HexColor('#FBF3E3')   # tint used behind filled cells
        border_grey = colors.HexColor('#D0DCE8')
        text_muted = colors.HexColor('#64748B')
        text_dark = colors.HexColor('#0F172A')
 
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            leftMargin=1.2 * cm, rightMargin=1.2 * cm,
            topMargin=1.4 * cm, bottomMargin=1.6 * cm,
        )
 
        # ---- Header block (letterhead-style) ----
        eyebrow_style = ParagraphStyle(
            'eyebrow', fontSize=9, fontName='Helvetica-Bold', alignment=1,
            textColor=brass, spaceAfter=2, tracking=1,
        )
        title_style = ParagraphStyle(
            'title', fontSize=18, fontName='Helvetica-Bold', alignment=1,
            textColor=navy, spaceAfter=2,
        )
        subtitle_style = ParagraphStyle(
            'subtitle', fontSize=10, fontName='Helvetica', alignment=1,
            textColor=text_muted, spaceAfter=10,
        )
 
        stream_label = 'Physical Science' if stream == 'physical' else 'Bio Science'
 
        eyebrow = Paragraph('FACULTY OF SCIENCE &nbsp;·&nbsp; UNIVERSITY OF RUHUNA', eyebrow_style)
        title = Paragraph(f"Level {level} Timetable — {stream_label}", title_style)
        subtitle = Paragraph(f"Semester {semester}", subtitle_style)
 
        # ---- Table ----
        header_style = ParagraphStyle('hdr', fontSize=9, fontName='Helvetica-Bold',
                                      alignment=1, textColor=colors.white)
        time_style = ParagraphStyle('t', fontSize=8.5, fontName='Helvetica-Bold',
                                    alignment=1, textColor=colors.white)
        course_style = ParagraphStyle('course', fontSize=8.5, fontName='Helvetica-Bold',
                                      alignment=1, textColor=navy, leading=11)
        venue_style = ParagraphStyle('venue', fontSize=7.5, fontName='Helvetica',
                                     alignment=1, textColor=brass, leading=10)
        lecturer_style = ParagraphStyle('lect', fontSize=7, fontName='Helvetica-Oblique',
                                        alignment=1, textColor=text_muted, leading=9)
 
        header_row = [Paragraph('Time', header_style)] + \
                     [Paragraph(d, header_style) for d in DAYS]
        rows = [header_row]
 
        for t in TIMES:
            row = [Paragraph(t, time_style)]
            for day in DAYS:
                slot = grid[day].get(t)
                if slot:
                    cell_content = [
                        Paragraph(slot.course.code, course_style),
                        Paragraph(slot.venue.code, venue_style),
                        Paragraph(slot.lecturer.name, lecturer_style),
                    ]
                    row.append(cell_content)
                else:
                    row.append('')
            rows.append(row)
 
        col_widths = [2.1 * cm] + [5.14 * cm] * 5
        tbl = Table(rows, colWidths=col_widths, rowHeights=[0.9 * cm] + [1.55 * cm] * len(TIMES))
 
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('BACKGROUND', (0, 1), (0, -1), navy_soft),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_grey),
            ('LINEBELOW', (0, 0), (-1, 0), 1.2, brass),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#F7F9FC')]),
        ]
        tbl.setStyle(TableStyle(table_style))
 
        # Tint filled cells with the brass-adjacent colour and a brass left-edge accent
        for r_idx, t in enumerate(TIMES, start=1):
            for c_idx, day in enumerate(DAYS, start=1):
                if grid[day].get(t):
                    tbl.setStyle(TableStyle([
                        ('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), brass_light),
                        ('LINEBEFORE', (c_idx, r_idx), (c_idx, r_idx), 2, brass),
                    ]))
 
        # ---- Footer (drawn on every page via canvas callback) ----
        generated_at = datetime.datetime.now().strftime('%d %b %Y, %H:%M')
 
        def draw_footer(canvas, doc_):
            canvas.saveState()
            canvas.setStrokeColor(border_grey)
            canvas.setLineWidth(0.5)
            canvas.line(1.2 * cm, 1.15 * cm, landscape(A4)[0] - 1.2 * cm, 1.15 * cm)
 
            canvas.setFont('Helvetica', 7.5)
            canvas.setFillColor(text_muted)
            canvas.drawString(1.2 * cm, 0.75 * cm,
                               f"Generated by Faculty Timetable Management System — {generated_at}")
            canvas.drawRightString(landscape(A4)[0] - 1.2 * cm, 0.75 * cm,
                                    f"Page {doc_.page}")
            canvas.restoreState()
 
        story = [eyebrow, title, subtitle, Spacer(1, 0.2 * cm), tbl]
        doc.build(story, onFirstPage=draw_footer, onLaterPages=draw_footer)
 
        buffer.seek(0)
        filename = f"timetable_level{level}_{stream}_{semester}.pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class LecturerScheduleViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ScheduleSlotReadSerializer

    def get_queryset(self):
        profile = getattr(self.request.user, 'profile', None)
        lecturer = profile.lecturer if profile else None
        if lecturer is None:
            return ScheduleSlot.objects.none()
        return ScheduleSlot.objects.select_related('timeslot', 'course', 'lecturer', 'venue', 'group').filter(lecturer=lecturer, is_published=True)

class LecturerRequestViewSet(viewsets.ModelViewSet):
    serializer_class = LecturerRequestSerializer

    def get_queryset(self):
        profile = getattr(self.request.user, 'profile', None)
        if profile and profile.role == 'ADMIN':
            return LecturerRequest.objects.select_related('lecturer', 'schedule_slot', 'reviewed_by').all()

        lecturer = profile.lecturer if profile else None
        if lecturer is None:
            return LecturerRequest.objects.none()
        return LecturerRequest.objects.select_related('lecturer', 'schedule_slot', 'reviewed_by').filter(lecturer=lecturer)

    def perform_create(self, serializer):
        profile = getattr(self.request.user, 'profile', None)
        lecturer = profile.lecturer if profile else None
        instance = serializer.save(lecturer=lecturer)

        request_type = self.request.data.get('request_type')
        if request_type == 'AVAILABILITY':
            student_group_ids = self.request.data.get('student_groups', [])
            if student_group_ids:
                lecturer_name = lecturer.name if lecturer else "A Lecturer"
                date_str = instance.requested_date.strftime('%Y-%m-%d') if instance.requested_date else "N/A"
                start_str = instance.requested_start.strftime('%H:%M') if instance.requested_start else "N/A"
                end_str = instance.requested_end.strftime('%H:%M') if instance.requested_end else "N/A"
                reason_str = instance.reason or "No reason supplied"

                title = f"Lecturer Availability: {lecturer_name}"
                message = f"Lecturer {lecturer_name} has indicated availability on {date_str} from {start_str} to {end_str}. Details: {reason_str}"

                for group_id in student_group_ids:
                    StudentNotification.objects.create(
                        student_group_id=group_id,
                        notification_type='GENERAL',
                        title=title,
                        message=message,
                    )

    @action(detail=True, methods=['post'], permission_classes=[IsAdminRole])
    def approve(self, request, pk=None):
        lecturer_request = self.get_object()
        lecturer_request.status = 'APPROVED'
        lecturer_request.reviewed_by = request.user
        lecturer_request.reviewed_at = timezone.now()
        lecturer_request.save()

        LecturerNotification.objects.create(
            lecturer=lecturer_request.lecturer,
            title="Request Approved",
            message=f"Your {lecturer_request.request_type.lower()} request has been approved.",
            notification_type='REQUEST',
            schedule_slot=lecturer_request.schedule_slot
        )

        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdminRole])
    def reject(self, request, pk=None):
        lecturer_request = self.get_object()
        lecturer_request.status = 'REJECTED'
        lecturer_request.reviewed_by = request.user
        lecturer_request.reviewed_at = timezone.now()
        lecturer_request.save()

        LecturerNotification.objects.create(
            lecturer=lecturer_request.lecturer,
            title="Request Rejected",
            message=f"Your {lecturer_request.request_type.lower()} request has been rejected.",
            notification_type='REQUEST',
            schedule_slot=lecturer_request.schedule_slot
        )

        return Response({'status': 'rejected'})


class LecturerNotificationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = LecturerNotificationSerializer

    def get_queryset(self):
        profile = getattr(self.request.user, 'profile', None)
        lecturer = profile.lecturer if profile else None
        if lecturer is None:
            return LecturerNotification.objects.none()
        return LecturerNotification.objects.select_related('lecturer', 'schedule_slot').filter(lecturer=lecturer)

    


# ============================================================
# Student dashboard / profile
# ============================================================

class StudentDashboardView(APIView):
    def get(self, request):
        profile = get_student_profile(request)
        if profile is None:
            return Response({'detail': 'Student profile not found'}, status=status.HTTP_403_FORBIDDEN)

        semester = request.query_params.get('semester', 'S2-2026')
        return Response(build_student_dashboard(profile, request, semester=semester))


class StudentProfileView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        profile = get_student_profile(request)
        if profile is None:
            return Response({'detail': 'Student profile not found'}, status=status.HTTP_403_FORBIDDEN)

        data = StudentProfileSerializer(profile, context={'request': request}).data
        curriculum_courses = profile.student_group.courses.all() if profile.student_group else Course.objects.none()
        data['enrolled_subjects'] = CourseSerializer(curriculum_courses, many=True).data
        return Response(data)

    def patch(self, request):
        profile = get_student_profile(request)
        if profile is None:
            return Response({'detail': 'Student profile not found'}, status=status.HTTP_403_FORBIDDEN)

        serializer = StudentProfileSerializer(profile, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()

        data = StudentProfileSerializer(profile, context={'request': request}).data
        curriculum_courses = profile.student_group.courses.all() if profile.student_group else Course.objects.none()
        data['enrolled_subjects'] = CourseSerializer(curriculum_courses, many=True).data
        return Response(data)


# ============================================================
# Admin — single create (Lecturer / Student)
# ============================================================

class AdminLecturerCreateView(APIView):
    permission_classes = [IsAdminRole]

    @transaction.atomic
    def post(self, request):
        name = request.data.get('name', '').strip()
        email = request.data.get('email', '').strip()
        department = request.data.get('department', '').strip()
        lecturer_id = request.data.get('lecturer_id', '').strip()
        password = request.data.get('password', '').strip()

        if not name or not email:
            return Response({'detail': 'Name and Email are required fields.'}, status=status.HTTP_400_BAD_REQUEST)

        if Lecturer.objects.filter(email=email).exists():
            return Response({'detail': 'Lecturer with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        if not lecturer_id:
            lecturer_id = generate_lecturer_id()
        elif Lecturer.objects.filter(lecturer_id=lecturer_id).exists():
            return Response({'detail': 'Lecturer ID must be unique.'}, status=status.HTTP_400_BAD_REQUEST)

        raw_password = password
        if not raw_password:
            raw_password = generate_random_password()

        lecturer = Lecturer.objects.create(
            lecturer_id=lecturer_id,
            name=name,
            email=email,
            department=department,
            must_change_password=True
        )

        parts = [part for part in name.split() if part]
        first_name = parts[0] if parts else ''
        last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

        if User.objects.filter(username=lecturer_id).exists():
            transaction.set_rollback(True)
            return Response({'detail': f'A user with username {lecturer_id} already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create(
            username=lecturer_id,
            email=email,
            first_name=first_name,
            last_name=last_name
        )
        user.set_password(raw_password)
        user.save()

        UserProfile.objects.create(
            user=user,
            role='LECTURER',
            lecturer=lecturer,
            must_change_password=True
        )

        send_credentials_email(name, email, lecturer_id, raw_password, 'Lecturer')

        return Response({
            'id': lecturer.id,
            'lecturer_id': lecturer_id,
            'name': name,
            'email': email,
            'department': department,
            'username': lecturer_id,
            'password': raw_password,
            'must_change_password': True
        }, status=status.HTTP_201_CREATED)


class AdminStudentCreateView(APIView):
    permission_classes = [IsAdminRole]

    @transaction.atomic
    def post(self, request):
        registration_number = request.data.get('registration_number', '').strip()
        name = request.data.get('name', '').strip()
        email = request.data.get('email', '').strip()
        student_group_id = request.data.get('student_group_id')
        contact_number = request.data.get('contact_number', '').strip()
        password = request.data.get('password', '').strip()

        if not registration_number or not name:
            return Response({'detail': 'Registration Number and Name are required fields.'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(username=registration_number).exists():
            return Response({'detail': 'A student account with this registration number/username already exists.'}, status=status.HTTP_400_BAD_REQUEST)
        if UserProfile.objects.filter(registration_number=registration_number).exists():
            return Response({'detail': 'Registration number is already in use.'}, status=status.HTTP_400_BAD_REQUEST)

        if not student_group_id:
            return Response({'detail': 'Student Group is required — it determines which subjects the student is enrolled in.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            student_group = StudentGroup.objects.get(pk=student_group_id)
        except StudentGroup.DoesNotExist:
            return Response({'detail': 'Invalid Student Group selected.'}, status=status.HTTP_400_BAD_REQUEST)

        raw_password = password
        if not raw_password:
            raw_password = generate_random_password()

        parts = [part for part in name.split() if part]
        first_name = parts[0] if parts else ''
        last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

        user = User.objects.create(
            username=registration_number,
            email=email,
            first_name=first_name,
            last_name=last_name
        )
        user.set_password(raw_password)
        user.save()

        UserProfile.objects.create(
            user=user,
            role='STUDENT',
            student_group=student_group,
            registration_number=registration_number,
            contact_number=contact_number,
            must_change_password=True
        )

        send_credentials_email(name, email, registration_number, raw_password, 'Student')

        return Response({
            'id': user.id,
            'username': registration_number,
            'registration_number': registration_number,
            'name': name,
            'email': email,
            'student_group_id': student_group_id,
            'contact_number': contact_number,
            'password': raw_password,
            'must_change_password': True
        }, status=status.HTTP_201_CREATED)


# ============================================================
# Admin — bulk upload (Student / Lecturer)
# ============================================================

class AdminBulkStudentUploadView(APIView):
    """
    POST /api/admin/students/bulk-upload/
    Accepts a .csv or .xlsx file (field name: 'file') with columns:
    name, registration_number, email, level, stream, year, subgroup (optional), contact_number (optional)
    """
    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'A file is required (field name: file).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fieldnames, data_rows = parse_uploaded_file(file)
        except Exception:
            return Response(
                {'detail': 'Could not read file. Please upload a .csv or .xlsx file matching the template.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        required_columns = {'name', 'registration_number', 'email', 'level', 'stream', 'year'}
        available_columns = set(fieldnames)

        if not required_columns.issubset(available_columns):
            missing = required_columns - available_columns
            return Response(
                {'detail': f'File is missing required columns: {", ".join(sorted(missing))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []

        for idx, row in enumerate(data_rows, start=2):
            name = (row.get('name') or '').strip()
            registration_number = (row.get('registration_number') or '').strip()
            email = (row.get('email') or '').strip()
            level = (row.get('level') or '').strip()
            stream = (row.get('stream') or '').strip()
            subgroup = (row.get('subgroup') or '').strip()
            year = (row.get('year') or '').strip()
            contact_number = (row.get('contact_number') or '').strip()

            try:
                with transaction.atomic():
                    if not name or not registration_number or not email:
                        raise ValueError('name, registration_number, and email are required.')

                    if User.objects.filter(username=registration_number).exists():
                        raise ValueError('Registration number already in use.')
                    if UserProfile.objects.filter(registration_number=registration_number).exists():
                        raise ValueError('Registration number already in use.')

                    try:
                        student_group = StudentGroup.objects.get(
                            level=level, stream=stream, subgroup=subgroup, year=year
                        )
                    except StudentGroup.DoesNotExist:
                        raise ValueError(
                            f'No matching student group for level={level}, stream={stream}, '
                            f'subgroup="{subgroup}", year={year}.'
                        )

                    raw_password = generate_random_password()
                    parts = [p for p in name.split() if p]
                    first_name = parts[0] if parts else ''
                    last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

                    user = User.objects.create(
                        username=registration_number, email=email,
                        first_name=first_name, last_name=last_name,
                    )
                    user.set_password(raw_password)
                    user.save()

                    UserProfile.objects.create(
                        user=user, role='STUDENT', student_group=student_group,
                        registration_number=registration_number, contact_number=contact_number,
                        must_change_password=True,
                    )

                send_credentials_email(name, email, registration_number, raw_password, 'Student')
                results.append({
                    'row': idx, 'status': 'success', 'name': name,
                    'registration_number': registration_number, 'email': email,
                })
            except Exception as exc:
                results.append({
                    'row': idx, 'status': 'error', 'name': name,
                    'registration_number': registration_number, 'detail': str(exc),
                })

        success_count = sum(1 for r in results if r['status'] == 'success')
        return Response({
            'total': len(results),
            'success_count': success_count,
            'failed_count': len(results) - success_count,
            'results': results,
        }, status=status.HTTP_201_CREATED)


class AdminBulkLecturerUploadView(APIView):
    """
    POST /api/admin/lecturers/bulk-upload/
    Accepts a .csv or .xlsx file (field name: 'file') with columns:
    name, email, department (optional), lecturer_id (optional — auto-generated if blank)
    """
    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'A file is required (field name: file).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fieldnames, data_rows = parse_uploaded_file(file)
        except Exception:
            return Response(
                {'detail': 'Could not read file. Please upload a .csv or .xlsx file matching the template.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # FIXED: this previously incorrectly required registration_number/
        # level/stream/year (copy-pasted from the Student upload view) —
        # fields that don't even exist on the Lecturer template.
        required_columns = {'name', 'email'}
        available_columns = set(fieldnames)

        if not required_columns.issubset(available_columns):
            missing = required_columns - available_columns
            return Response(
                {'detail': f'File is missing required columns: {", ".join(sorted(missing))}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []

        for idx, row in enumerate(data_rows, start=2):
            name = (row.get('name') or '').strip()
            email = (row.get('email') or '').strip()
            department = (row.get('department') or '').strip()
            lecturer_id = (row.get('lecturer_id') or '').strip()

            try:
                with transaction.atomic():
                    if not name or not email:
                        raise ValueError('name and email are required.')

                    if Lecturer.objects.filter(email=email).exists():
                        raise ValueError('Lecturer with this email already exists.')

                    if not lecturer_id:
                        lecturer_id = generate_lecturer_id()
                    elif Lecturer.objects.filter(lecturer_id=lecturer_id).exists():
                        raise ValueError('Lecturer ID must be unique.')

                    if User.objects.filter(username=lecturer_id).exists():
                        raise ValueError(f'A user with username {lecturer_id} already exists.')

                    raw_password = generate_random_password()

                    lecturer = Lecturer.objects.create(
                        lecturer_id=lecturer_id, name=name, email=email,
                        department=department, must_change_password=True,
                    )

                    parts = [p for p in name.split() if p]
                    first_name = parts[0] if parts else ''
                    last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

                    user = User.objects.create(
                        username=lecturer_id, email=email,
                        first_name=first_name, last_name=last_name,
                    )
                    user.set_password(raw_password)
                    user.save()

                    UserProfile.objects.create(
                        user=user, role='LECTURER', lecturer=lecturer, must_change_password=True,
                    )

                send_credentials_email(name, email, lecturer_id, raw_password, 'Lecturer')
                results.append({
                    'row': idx, 'status': 'success', 'name': name,
                    'lecturer_id': lecturer_id, 'email': email,
                })
            except Exception as exc:
                results.append({
                    'row': idx, 'status': 'error', 'name': name,
                    'lecturer_id': lecturer_id, 'detail': str(exc),
                })

        success_count = sum(1 for r in results if r['status'] == 'success')
        return Response({
            'total': len(results),
            'success_count': success_count,
            'failed_count': len(results) - success_count,
            'results': results,
        }, status=status.HTTP_201_CREATED)


# ============================================================
# Admin — bulk upload templates (.xlsx)
# ============================================================

def _styled_header(ws, headers, col_widths):
    """Shared styling so both templates look consistent with the rest of your app's navy theme."""
    ws.append(headers)
    header_fill = PatternFill(start_color='0D1B2A', end_color='0D1B2A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(style='thin', color='D0DCE8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'
    return border


class AdminStudentBulkTemplateView(APIView):
    """GET /api/admin/students/bulk-template/ — downloads a styled .xlsx template."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        headers = ['name', 'registration_number', 'email', 'level', 'stream', 'year', 'subgroup', 'contact_number']
        widths = [22, 20, 26, 8, 12, 8, 10, 16]
        border = _styled_header(ws, headers, widths)

        sample = ['W.M. Perera', 'SC/2022/1023', 'perera@example.com', 'I', 'physical', '2024', '', '0771234567']
        ws.append(sample)
        for cell in ws[2]:
            cell.font = Font(italic=True, color='94A3B8')
            cell.border = border

        level_dv = DataValidation(type='list', formula1='"I,II,III"', allow_blank=True,
                                   errorTitle='Invalid Level', error='Must be I, II, or III')
        stream_dv = DataValidation(type='list', formula1='"physical,bio,both"', allow_blank=True,
                                    errorTitle='Invalid Stream', error='Must be physical, bio, or both')
        ws.add_data_validation(level_dv)
        ws.add_data_validation(stream_dv)
        level_dv.add('D2:D500')
        stream_dv.add('E2:E500')

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="student_bulk_template.xlsx"'
        return response


class AdminLecturerBulkTemplateView(APIView):
    """GET /api/admin/lecturers/bulk-template/ — downloads a styled .xlsx template."""
    permission_classes = [IsAdminRole]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lecturers'

        headers = ['name', 'email', 'department', 'lecturer_id']
        widths = [24, 26, 22, 18]
        border = _styled_header(ws, headers, widths)

        sample = ['Dr. A.B. Silva', 'silva@example.com', 'Computer Science', '']
        ws.append(sample)
        for cell in ws[2]:
            cell.font = Font(italic=True, color='94A3B8')
            cell.border = border

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="lecturer_bulk_template.xlsx"'
        return response


# ============================================================
# Admin — analytics / availability / auto-scheduler / curriculum
# ============================================================

# Replace your existing AdminFreeSlotsView with this version.

class AdminFreeSlotsView(APIView):
    """
    GET /api/admin/analytics/free-slots/?type=venue&id=<id>&semester=<semester>
    GET /api/admin/analytics/free-slots/?type=lecturer&id=<id>&semester=<semester>

    Returns every TimeSlot in the system, each tagged with is_free.
    For booked slots, also returns `occupants` — who/what is actually
    using that slot, so the frontend can show it on hover:
      - venue mode:    course, lecturer name, group(s) in that room
      - lecturer mode:  course, venue code, group(s) that lecturer is teaching
    A list rather than a single object, since a joint/combined class can
    have multiple student groups sharing the same slot.
    """
    permission_classes = [IsAdminRole]

    def get(self, request):
        target_type = request.query_params.get('type', 'venue').strip().lower()
        target_id = request.query_params.get('id')
        semester = request.query_params.get('semester')

        if not target_id:
            return Response({'detail': 'id query param is required'}, status=status.HTTP_400_BAD_REQUEST)

        if target_type not in ('venue', 'lecturer'):
            return Response({'detail': 'type must be "venue" or "lecturer"'}, status=status.HTTP_400_BAD_REQUEST)

        all_slots = TimeSlot.objects.all().order_by('day', 'start_time')

        schedule_qs = ScheduleSlot.objects.select_related('course', 'lecturer', 'venue', 'group').all()
        if target_type == 'lecturer':
            schedule_qs = schedule_qs.filter(lecturer_id=target_id)
        else:
            schedule_qs = schedule_qs.filter(venue_id=target_id)

        if semester:
            schedule_qs = schedule_qs.filter(semester=semester)

        # Build timeslot_id -> list of occupant dicts
        occupant_map = {}
        for row in schedule_qs:
            occupant_info = {
                'course_code': row.course.code,
                'course_name': row.course.name,
                'group_display': str(row.group),
            }
            if target_type == 'venue':
                occupant_info['lecturer_name'] = row.lecturer.name
            else:
                occupant_info['venue_code'] = row.venue.code

            occupant_map.setdefault(row.timeslot_id, []).append(occupant_info)

        results = [
            {
                'id': slot.id,
                'day': slot.day,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
                'is_free': slot.id not in occupant_map,
                'occupants': occupant_map.get(slot.id, []),
            }
            for slot in all_slots
        ]

        return Response(results)


class AdminAnalyticsSummaryView(APIView):
    """GET /api/admin/analytics/summary/?semester=<semester>"""
    permission_classes = [IsAdminRole]

    def get(self, request):
        semester = request.query_params.get('semester', 'S2-2026')
        all_timeslots_count = TimeSlot.objects.count() or 1

        room_utilization = []
        for venue in Venue.objects.all():
            booked = (
                ScheduleSlot.objects
                .filter(venue=venue, semester=semester)
                .values('timeslot_id')
                .distinct()
                .count()
            )
            room_utilization.append({
                'venue_code': venue.code,
                'venue_name': venue.name,
                'booked_slots': booked,
                'total_slots': all_timeslots_count,
                'utilization_pct': round((booked / all_timeslots_count) * 100, 1),
            })
        room_utilization.sort(key=lambda item: -item['utilization_pct'])

        lecturer_workload = []
        for lecturer in Lecturer.objects.all():
            slots = list(
                ScheduleSlot.objects
                .select_related('timeslot')
                .filter(lecturer=lecturer, semester=semester)
            )
            total_minutes = 0
            for slot in slots:
                start = slot.timeslot.start_time
                end = slot.timeslot.end_time
                total_minutes += (end.hour * 60 + end.minute) - (start.hour * 60 + start.minute)
            lecturer_workload.append({
                'lecturer_name': lecturer.name,
                'classes_count': len(slots),
                'weekly_hours': round(total_minutes / 60, 1),
            })
        lecturer_workload.sort(key=lambda item: -item['weekly_hours'])

        day_counts = {day: 0 for day in DAYS}
        for slot in ScheduleSlot.objects.select_related('timeslot').filter(semester=semester):
            day_counts[slot.timeslot.day] = day_counts.get(slot.timeslot.day, 0) + 1
        day_distribution = [{'day': day, 'classes': day_counts.get(day, 0)} for day in DAYS]

        time_counter = Counter()
        for slot in ScheduleSlot.objects.select_related('timeslot').filter(semester=semester):
            key = (slot.timeslot.day, slot.timeslot.start_time.strftime('%H:%M'))
            time_counter[key] += 1

        busiest_times = [
            {'day': key[0], 'time': key[1], 'count': count}
            for key, count in sorted(time_counter.items(), key=lambda kv: -kv[1])[:8]
        ]

        return Response({
            'semester': semester,
            'room_utilization': room_utilization,
            'lecturer_workload': lecturer_workload,
            'day_distribution': day_distribution,
            'busiest_times': busiest_times,
        })


class AdminAutoScheduleView(APIView):
    """POST /api/admin/scheduling/auto-generate/"""
    permission_classes = [IsAdminRole]

    def post(self, request):
        group_id = request.data.get('group_id')
        semester = request.data.get('semester', 'S2-2026')
        requirements = request.data.get('requirements', [])

        if not group_id:
            return Response({'detail': 'group_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not requirements:
            return Response({'detail': 'At least one requirement (course + lecturer) is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            group = StudentGroup.objects.get(pk=group_id)
        except StudentGroup.DoesNotExist:
            return Response({'detail': 'Student group not found.'}, status=status.HTTP_404_NOT_FOUND)

        resolved_requirements = []
        for req in requirements:
            try:
                course = Course.objects.get(pk=req.get('course_id'))
                lecturer = Lecturer.objects.get(pk=req.get('lecturer_id'))
            except (Course.DoesNotExist, Lecturer.DoesNotExist):
                return Response(
                    {'detail': f'Invalid course_id or lecturer_id in requirement: {req}'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            resolved_requirements.append({
                'course': course,
                'lecturer': lecturer,
                'lecturer_id': lecturer.id,
                'venue_type': req.get('venue_type'),
            })

        assignments, is_complete = generate_timetable_for_group(group_id, semester, resolved_requirements)

        results = []
        for req, assignment in zip(resolved_requirements, assignments):
            if assignment is None:
                results.append({
                    'course_id': req['course'].id,
                    'course_code': req['course'].code,
                    'course_name': req['course'].name,
                    'lecturer_id': req['lecturer'].id,
                    'lecturer_name': req['lecturer'].name,
                    'status': 'unassigned',
                })
            else:
                timeslot, venue = assignment
                results.append({
                    'course_id': req['course'].id,
                    'course_code': req['course'].code,
                    'course_name': req['course'].name,
                    'lecturer_id': req['lecturer'].id,
                    'lecturer_name': req['lecturer'].name,
                    'timeslot_id': timeslot.id,
                    'day': timeslot.day,
                    'start_time': timeslot.start_time.strftime('%H:%M'),
                    'end_time': timeslot.end_time.strftime('%H:%M'),
                    'venue_id': venue.id,
                    'venue_code': venue.code,
                    'venue_name': venue.name,
                    'status': 'assigned',
                })

        return Response({
            'group_id': group_id,
            'group_display': str(group),
            'semester': semester,
            'is_complete': is_complete,
            'assigned_count': sum(1 for r in results if r['status'] == 'assigned'),
            'unassigned_count': sum(1 for r in results if r['status'] == 'unassigned'),
            'results': results,
        })


class AdminCurriculumView(APIView):
    """
    GET /api/admin/curriculum/?group_id=<id>  -> {group_id, course_ids: [...]}
    PUT /api/admin/curriculum/                -> body: {group_id, course_ids: [...]}
    """
    permission_classes = [IsAdminRole]

    def get(self, request):
        group_id = request.query_params.get('group_id')
        if not group_id:
            return Response({'detail': 'group_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            group = StudentGroup.objects.get(pk=group_id)
        except StudentGroup.DoesNotExist:
            return Response({'detail': 'Student group not found.'}, status=status.HTTP_404_NOT_FOUND)

        return Response({
            'group_id': group.id,
            'course_ids': list(group.courses.values_list('id', flat=True)),
        })

    def put(self, request):
        group_id = request.data.get('group_id')
        course_ids = request.data.get('course_ids', [])

        if not group_id:
            return Response({'detail': 'group_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            group = StudentGroup.objects.get(pk=group_id)
        except StudentGroup.DoesNotExist:
            return Response({'detail': 'Student group not found.'}, status=status.HTTP_404_NOT_FOUND)

        courses_qs = Course.objects.filter(id__in=course_ids)
        group.courses.set(courses_qs)

        return Response({
            'group_id': group.id,
            'course_ids': list(group.courses.values_list('id', flat=True)),
        })


# ============================================================
# Admin — admin account management
# ============================================================

class AdminAccountsView(APIView):
    """
    GET  /api/admin/admins/  -> list every Admin account
    POST /api/admin/admins/  -> create a new Admin account
    """
    permission_classes = [IsAdminRole]

    def get(self, request):
        admins = User.objects.filter(profile__role='ADMIN').order_by('username')
        results = [
            {
                'id': admin.id,
                'username': admin.username,
                'name': admin.get_full_name().strip() or admin.username,
                'email': admin.email,
                'date_joined': admin.date_joined,
                'is_you': admin.id == request.user.id,
            }
            for admin in admins
        ]
        return Response(results)

    @transaction.atomic
    def post(self, request):
        name = request.data.get('name', '').strip()
        email = request.data.get('email', '').strip()
        password = request.data.get('password', '').strip()

        if not name or not email:
            return Response({'detail': 'Name and Email are required fields.'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email=email).exists():
            return Response({'detail': 'An account with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(username=email).exists():
            return Response({'detail': 'An account with this username already exists.'}, status=status.HTTP_400_BAD_REQUEST)

        raw_password = password or generate_random_password()

        parts = [p for p in name.split() if p]
        first_name = parts[0] if parts else ''
        last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

        user = User.objects.create(
            username=email, email=email,
            first_name=first_name, last_name=last_name,
        )
        user.set_password(raw_password)
        user.save()

        UserProfile.objects.create(user=user, role='ADMIN', must_change_password=True)

        send_credentials_email(name, email, email, raw_password, 'Admin')

        return Response({
            'id': user.id,
            'username': email,
            'name': name,
            'email': email,
            'password': raw_password,
            'must_change_password': True,
        }, status=status.HTTP_201_CREATED)


class AdminMeView(APIView):
    """
    GET   /api/admin/me/  -> current admin's own profile
    PATCH /api/admin/me/  -> update own name, email, avatar (multipart/form-data)
    """
    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        return Response(serialize_user(request.user, request=request))

    def patch(self, request):
        user = request.user
        name = request.data.get('name', '').strip()
        email = request.data.get('email', '').strip()

        if name:
            parts = [p for p in name.split() if p]
            user.first_name = parts[0] if parts else ''
            user.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''
        if email:
            if User.objects.filter(email=email).exclude(id=user.id).exists():
                return Response({'detail': 'This email is already in use by another account.'}, status=status.HTTP_400_BAD_REQUEST)
            user.email = email
        user.save()

        avatar_file = request.FILES.get('avatar')
        if avatar_file:
            profile = getattr(user, 'profile', None)
            if profile:
                profile.avatar = avatar_file
                profile.save(update_fields=['avatar'])

        return Response(serialize_user(user, request=request))


class AdminTimetablePublishView(APIView):
    """
    POST /api/admin/timetable/publish/
    Body: { "level": "I", "stream": "physical", "semester": "S2-2026", "action": "publish" | "unpublish" }
 
    Bulk-flips is_published for every ScheduleSlot matching the given
    level/stream/semester — this is the moment a draft timetable becomes
    visible to Student/Lecturer dashboards (or gets pulled back to draft).
    """
    permission_classes = [IsAdminRole]
 
    def post(self, request):
        level = request.data.get('level')
        stream = request.data.get('stream')
        semester = request.data.get('semester', 'S2-2026')
        action = request.data.get('action', 'publish')
 
        if not level or not stream:
            return Response({'detail': 'level and stream are required.'}, status=status.HTTP_400_BAD_REQUEST)
 
        if action not in ('publish', 'unpublish'):
            return Response({'detail': 'action must be "publish" or "unpublish".'}, status=status.HTTP_400_BAD_REQUEST)
 
        publish_state = action == 'publish'
 
        qs = ScheduleSlot.objects.filter(group__level=level, group__stream=stream, semester=semester)
        updated_count = qs.update(is_published=publish_state)
 
        return Response({
            'level': level,
            'stream': stream,
            'semester': semester,
            'action': action,
            'updated_count': updated_count,
        })